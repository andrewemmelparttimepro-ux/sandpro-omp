#!/usr/bin/env python3
"""Regenerate src/data/okr2026Consolidated.json from the consolidated OKR sheet.

Source: "2026 All Group OKRs_ Consolidated_4.8.26.xlsx" ("MAIN OKR SHEET"),
delivered by Mercileidy Jimenez after the Tim/Jake meeting.

Usage:
    python3 scripts/gen-okr-2026.py [path/to/2026 All Group OKRs_ Consolidated.xlsx]

Defaults to the copy in ~/Downloads/ompframeworkokrsreferencedocuments/.
Requires: openpyxl  (pip install openpyxl)

Structure produced: 1 COMPANY top-line group + 17 group scorecards.
Header rows are detected by column C == "AUDIT FORM"; "ACTUALS" marker rows are
skipped; data rows carry verbatim baseline/target/owner/status/cadence + Jan-Dec
actuals (columns J..U).
"""
import json
import os
import sys

import openpyxl

DEFAULT_SRC = os.path.expanduser(
    "~/Downloads/ompframeworkokrsreferencedocuments/2026 All Group OKRs_ Consolidated_4.8.26.xlsx"
)
OUT = os.path.join(os.path.dirname(__file__), "..", "src", "data", "okr2026Consolidated.json")
MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["MAIN OKR SHEET"]

    groups, cur = [], None
    for r in range(1, ws.max_row + 1):
        b, c = cell(ws, r, 2), cell(ws, r, 3)
        j = cell(ws, r, 10)
        if isinstance(j, str) and j.strip().upper() == "ACTUALS":
            continue
        if isinstance(c, str) and c.strip().upper() == "AUDIT FORM":
            cur = {"group": b, "isCompany": bool(b and "COMPANY" in b.upper()), "okrs": []}
            groups.append(cur)
            continue
        if cur and b:
            actuals = {}
            for i, m in enumerate(MONTHS):
                val = cell(ws, r, 10 + i)
                if val is not None:
                    actuals[m] = val
            cur["okrs"].append({
                "title": b,
                "auditForm": (str(c).strip().upper() == "Y") if c is not None else False,
                "baseline": cell(ws, r, 4),
                "target": cell(ws, r, 5),
                "owner": cell(ws, r, 6),
                "manualStatus": cell(ws, r, 7),
                "reportingCadence": cell(ws, r, 8),
                "ytdAvg": cell(ws, r, 9),
                "actuals": actuals,
            })

    data = {
        "source": os.path.basename(src),
        "sheet": "MAIN OKR SHEET",
        "groups": groups,
    }
    with open(OUT, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)

    total = sum(len(g["okrs"]) for g in groups)
    non_company = sum(1 for g in groups if not g["isCompany"])
    print(f"Wrote {os.path.relpath(OUT)}: {len(groups)} groups "
          f"(1 company + {non_company}), {total} OKRs")


if __name__ == "__main__":
    main()
